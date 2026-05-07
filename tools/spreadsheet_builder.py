"""Generate .xlsx spreadsheets from a niche/topic prompt.

Claude designs the structure (sheets, columns, formulas, sample data)
returning a strict JSON schema. openpyxl then materializes it.
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import anthropic
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .. import config

DESIGN_SYSTEM_PROMPT = """You design Excel/Google-Sheets spreadsheets that solve real operator problems.

You produce ONLY valid JSON matching this schema:

{
  "title": "string — the spreadsheet's user-facing name",
  "sheets": [
    {
      "name": "string — sheet tab name (max 31 chars)",
      "description": "string — one-sentence purpose",
      "columns": [
        {
          "header": "string",
          "key": "string — short snake_case for formula refs",
          "width": int,
          "format": "general|currency|percent|date|number|text"
        }
      ],
      "rows": [
        {"<key>": "literal value or =FORMULA"}
      ],
      "notes": ["string — usage tips shown in a Notes section below the table"]
    }
  ],
  "instructions_sheet": {
    "title": "How to use",
    "sections": [
      {"heading": "string", "body": "string (multi-line ok)"}
    ]
  }
}

Rules:
- Use real Excel formulas (=SUM(B2:B10), =IF(...), =VLOOKUP(...)) — they MUST be syntactically valid.
- Cell references use A1 notation. Absolute refs ($A$1) where appropriate.
- Provide 3-10 sample rows so the buyer sees how it works.
- Include an instructions_sheet so non-technical users can self-serve.
- For pricing/cost tools, include built-in benchmarks where common (e.g., 30% food cost target).
- Aim for genuine usefulness. No filler columns."""


def _clean_json(text: str) -> dict:
    """Strip code fences and parse JSON Claude returned."""
    text = text.strip()
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text)
    return json.loads(text)


def design_spreadsheet(brief: str) -> dict:
    """Ask Claude to design the spreadsheet structure for the given brief."""
    client = anthropic.Anthropic(api_key=config.ANTHROPIC_API_KEY)
    response = client.messages.create(
        model=config.MODEL,
        max_tokens=16000,
        thinking={"type": "adaptive"},
        system=[{
            "type": "text",
            "text": DESIGN_SYSTEM_PROMPT,
            "cache_control": {"type": "ephemeral"},
        }],
        messages=[{"role": "user", "content": brief}],
    )
    text = next((b.text for b in response.content if b.type == "text"), "")
    return _clean_json(text)


# ---------- Materialization ----------

NUMBER_FORMATS = {
    "general": "General",
    "currency": '"$"#,##0.00',
    "percent": "0.00%",
    "date": "yyyy-mm-dd",
    "number": "#,##0.00",
    "text": "@",
}

HEADER_FILL = PatternFill(start_color="2F4858", end_color="2F4858", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
NOTE_FONT = Font(italic=True, color="555555", size=10)


def _write_sheet(wb: Workbook, sheet_def: dict) -> None:
    name = sheet_def["name"][:31]
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        ws = wb.active
        ws.title = name
    else:
        ws = wb.create_sheet(name)

    columns = sheet_def["columns"]
    keys = [c["key"] for c in columns]

    # Header row
    for i, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=i, value=col["header"])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = col.get("width", 18)

    # Data rows
    for r, row in enumerate(sheet_def.get("rows", []), start=2):
        for i, key in enumerate(keys, start=1):
            value = row.get(key, "")
            cell = ws.cell(row=r, column=i, value=value)
            fmt_name = columns[i - 1].get("format", "general")
            cell.number_format = NUMBER_FORMATS.get(fmt_name, "General")

    # Notes block below the table
    notes = sheet_def.get("notes") or []
    if notes:
        start_row = ws.max_row + 2
        ws.cell(row=start_row, column=1, value="Notes:").font = Font(bold=True)
        for j, note in enumerate(notes, start=1):
            ws.cell(row=start_row + j, column=1, value=f"• {note}").font = NOTE_FONT

    ws.freeze_panes = "A2"


def _write_instructions(wb: Workbook, instructions: dict) -> None:
    ws = wb.create_sheet(instructions.get("title", "How to use")[:31], 0)
    ws.column_dimensions["A"].width = 100
    ws.cell(row=1, column=1, value=instructions.get("title", "How to use")).font = Font(
        bold=True, size=16
    )
    row = 3
    for section in instructions.get("sections", []):
        ws.cell(row=row, column=1, value=section["heading"]).font = Font(
            bold=True, size=12
        )
        row += 1
        for line in section["body"].split("\n"):
            ws.cell(row=row, column=1, value=line).alignment = Alignment(wrap_text=True)
            row += 1
        row += 1


def build_xlsx(design: dict, output_path: str | Path) -> Path:
    """Materialize a design dict into an .xlsx file. Returns the path written."""
    wb = Workbook()
    # Default sheet gets renamed by _write_sheet on the first sheet
    for sheet_def in design["sheets"]:
        _write_sheet(wb, sheet_def)
    if "instructions_sheet" in design:
        _write_instructions(wb, design["instructions_sheet"])

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def design_and_build(brief: str, output_dir: str | Path) -> dict[str, Any]:
    """End-to-end: brief → design → .xlsx file. Returns design + file path."""
    design = design_spreadsheet(brief)
    safe_title = re.sub(r"[^A-Za-z0-9_-]+", "_", design["title"]).strip("_")[:60]
    out_path = Path(output_dir) / f"{safe_title}.xlsx"
    build_xlsx(design, out_path)
    return {"design": design, "file_path": str(out_path), "title": design["title"]}
